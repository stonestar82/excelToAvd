from hashlib import new
import re
from datetime import datetime
from src.process.DHCPScan import DHCPScan
from src.process.SwitchIdentification import SwitchIdentification
from src.domain.Switch import Switch
from operator import eq, ne
from openpyxl import load_workbook

def taskPrint(task):
  task = task + " "
  print(task.ljust(100, "*") + "\n")


##### dhcp 추출 S #####
taskPrint("TASK [Start]")
taskPrint("TASK [DHCP scan]")
with open("/workspace/excelToAvd/dhcp.leases", "r") as f:
  now = datetime.now()
  spineCount = 2
  leafCount = 4  
  scanSwitches = []
  dhcpScan = DHCPScan()
  
  data = f.read().split("lease")
  for line in data:
    ## ip 와 mac이 추출되면 end time 추출 및 현재 시간과 비교하여 유효분만 필터링
    ip = dhcpScan.searchIp(line)  
    mac = dhcpScan.searchMac(line)
    
    if ne("", ip) and ne("", mac):
      info = line.split(";")
      ## info[1] end time
      t = dhcpScan.searchTime(info[1])

      if ne("", t):
        endTime = datetime.strptime(t, "%Y/%m/%d %H:%M:%S")       
        dateDiff = now < endTime
        
        if dateDiff:
          obj = {"ip": ip, "mac": mac, "time":t}
          scanSwitches.append(obj)

# print(scanSwitches)
##### dhcp 추출 E #####


##### lldp 로 spine leaf 체크 S #####
taskPrint("TASK [LLDP scan]")
totalSpineCount = 2
totalLeafCount = 4

for item in scanSwitches:
    
  switch = Switch(item.get("ip"), "50:fa:80:e9:68:cf", "arista_eos")

  switch.user = "ansible"
  switch.password = "ansible"

  switchId = SwitchIdentification(switch)
  switchId.switchConnect()
  switchId.lldpScan()
  switchId.identification(totalLeafCount)
  switchId.disconnect()
  item.update(dict([("hostname", switchId.switch.hostname)]))

##### lldp 로 spine leaf 체크 E #####


##### excel 정보 업데이트 S #####
excel = load_workbook(filename="/workspace/excelToAvd/inventory.xlsx", read_only=False, data_only=True)

spineSheet = excel["Spine Info"]
leafSheet = excel["L3 Leaf Info"]

## spine sheet에 ip 정보 갱신
for i in range(2,spineSheet.max_row + 1):
  # print(spineSheet.cell(i, 2).value)
  for switch in scanSwitches:
    # print("hostname = ", switch.get("hostname"), spineSheet.cell(i, 2).value)
    if eq(switch.get("hostname"), spineSheet.cell(i, 2).value):
      spineSheet.cell(i, 3, switch.get("ip") + "/24")

## leaf sheet에 ip 정보 갱신
for i in range(2,leafSheet.max_row + 1):
  # print(leafSheet.cell(i, 4).value)
  for switch in scanSwitches:
    if eq(switch.get("hostname"), leafSheet.cell(i, 4).value):
      leafSheet.cell(i, 5, switch.get("ip") + "/24")   

excel.save("/workspace/excelToAvd/inventory.xlsx")
##### excel 정보 업데이트 E #####

taskPrint("TASK [End]")