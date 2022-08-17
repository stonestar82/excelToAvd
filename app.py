from flask import Flask, render_template, request, Response
import os, json
from operator import eq
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

app = Flask(__name__, template_folder="static/templates")

import logging
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

@app.route("/")
def index():

  switchList = os.listdir("./inventory/intended/configs/", )
  
  return render_template('switchList.html', switchList=switchList)

@app.route("/switch/<switch>")
def read(switch): 

  try:
    with open('./inventory/intended/configs/' + secure_filename(switch) + '.cfg', encoding="UTF-8") as f:
      cfg = f.readlines()
  except FileNotFoundError:
      cfg = ["설정파일이 없습니다."]

  return render_template('cfg.html', cfg=cfg)

@app.route("/cfgs/<switch>")
def cfgs(switch):
  try:
    with open('./inventory/intended/configs/' + secure_filename(switch) + '.cfg', encoding="UTF-8") as f:
      cfg = f.readlines()
      return "\n".join(cfg)
  except FileNotFoundError:
      return Response("", status=404, mimetype='application/json')

  

@app.route("/cfg/input")
def configInput():
  return render_template('configInput.html')

@app.route("/cfg/upload", methods = ['POST'])
def configUpload(): 

  f = request.files["cfg_file"]
  f.save("./upload/" + secure_filename(f.filename))

  return "파일이 저장되었습니다."

  
@app.route("/bootstrap")
def bootstrap():
  
  boot = ""
  with open('./bootstrap', encoding="utf-8") as f:
    boot = f.readlines()
  
  return "\n".join(boot)

@app.route("/bootstrap/status")
def requestStatus():
  
  excel = load_workbook(filename="./ip.xlsx", read_only=False, data_only=True)
  sheetIp = excel["ip"]
  
  status = []
  for row in range(2, sheetIp.max_row + 1):
    v, s, mac, ser = sheetIp.cell(row,1).value, sheetIp.cell(row,2).value, sheetIp.cell(row,3).value, sheetIp.cell(row,4).value
    d = { "ip": v, "status": s, "mac": mac, "serial": ser}
    status.append(d)  
    
  return render_template('bootstrapStatus.html', status=status)

@app.route("/bootstrap/requestip/<bootseq>/<sysmac>/<serial>")
def requestip(bootseq, sysmac, serial):
  
  excel = load_workbook(filename="./ip.xlsx", read_only=False, data_only=True)
  sheetIp = excel["ip"]

  ip = ""
  ipExsist = False
  for row in range(2, sheetIp.max_row + 1):
    v, s = sheetIp.cell(row,1).value, sheetIp.cell(row,2).value
    if eq("X", s):
      ip = v
      ipExsist = True
      sheetIp.cell(row, 2, "O")
      sheetIp.cell(row, 3, sysmac)
      sheetIp.cell(row, 4, serial)
      excel.save("./ip.xlsx")
      break;
    
  if ipExsist:
    
    if not os.path.exists("./upload/" + bootseq):
      with open("./upload/" + bootseq, "w") as f:
        f.write(ip + "|" + sysmac+ "|" + serial  + "\n")
        f.close()            
    else:
      with open("./upload/" + bootseq, "at", encoding="utf-8") as f:
        f.write(ip + "|" + sysmac + "|" + serial + "\n")
        f.close()
        
  else:
    ip = ""
    
  return ip


if __name__ == "__main__":
  app.run(host='0.0.0.0', port=80)